import csv
import io
import json
import logging
import re
import zipfile
from uuid import UUID
from urllib.parse import urlparse

import requests
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import Person, Company

log = logging.getLogger(__name__)

router = APIRouter()

# (db_field, display_header) — order matches desired output
EXPORT_COLUMNS = [
    ("image_url", "Photo"),
    ("name", "Name"),
    ("title", "Title"),
    ("_company_name", "Company / Practice"),
    ("linkedin_url", "LinkedIn URL"),
    ("linkedin_headline", "LinkedIn Headline"),
    ("primary_expertise", "Primary Expertise"),
    ("justification", "Justification (LinkedIn + Bio/Website)"),
    ("matched_13_categories", "Matched 13 Expertise Categories"),
    ("sector", "Sector"),
    ("geography", "Geography"),
    ("inferred_expertise_functional", "Inferred Expertise (Functional)"),
    ("matched_inferred_expertise_topics", "Matched Inferred Expertise (Topics)"),
    ("linkedin_experience_summary", "LinkedIn Experience Summary"),
    ("data_source", "Data Source"),
]

EXPORT_HEADERS = [h for _, h in EXPORT_COLUMNS]

# Column widths matching reference layout
COLUMN_WIDTHS = [18, 24, 22, 38, 40, 40, 28, 65, 48, 40, 35, 65, 70, 60, 22]


def _people_to_dicts(people, company_name: str = "") -> list[dict]:
    rows = []
    for p in people:
        row = {}
        for field, header in EXPORT_COLUMNS:
            if field == "_company_name":
                val = company_name
            else:
                val = getattr(p, field, None)
            if isinstance(val, list):
                val = "; ".join(str(v) for v in val if v)
            row[header] = val if val else "—"
        rows.append(row)
    return rows


def _sanitize_filename(name: str) -> str:
    """Remove characters not safe for filenames."""
    return re.sub(r'[<>:"/\\|?*]', '', name).strip()


def _download_photo(url: str, timeout: int = 10) -> bytes | None:
    """Download a photo, return bytes or None on failure."""
    try:
        resp = requests.get(url, timeout=timeout, headers={"Referer": ""})
        if resp.status_code == 200 and len(resp.content) > 500:
            return resp.content
    except Exception:
        pass
    return None


def _guess_extension(url: str, content: bytes) -> str:
    """Guess image extension from URL or content."""
    path = urlparse(url).path.lower()
    if path.endswith(".png"):
        return ".png"
    if path.endswith(".webp"):
        return ".webp"
    # Check magic bytes
    if content[:4] == b'\x89PNG':
        return ".png"
    if content[:4] == b'RIFF' and content[8:12] == b'WEBP':
        return ".webp"
    return ".jpg"


def _style_worksheet(ws):
    """Apply column widths and header styling to match reference layout."""
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    # Column widths
    for i, width in enumerate(COLUMN_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Bold headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)

    # Wrap text for long content columns
    wrap_cols = {7, 8, 9, 12, 13, 14}  # Justification, Categories, etc.
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")


@router.get("/export/{company_id}")
def export_company(
    company_id: UUID,
    format: str = Query("csv", pattern="^(csv|json|xlsx|zip)$"),
    db: Session = Depends(get_db),
):
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    people = db.query(Person).filter(Person.company_id == company_id).order_by(Person.name).all()
    company_name = company.name or ""
    rows = _people_to_dicts(people, company_name)
    safe_name = (company_name or "export").replace(" ", "_").lower()

    if format == "csv":
        output = io.StringIO()
        if rows:
            writer = csv.DictWriter(output, fieldnames=EXPORT_HEADERS)
            writer.writeheader()
            writer.writerows(rows)
        content = output.getvalue()
        return StreamingResponse(
            iter([content]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}_profiles.csv"'},
        )

    elif format == "json":
        content = json.dumps({
            "company": {"name": company_name, "url": company.url},
            "people_count": len(rows),
            "people": rows,
        }, ensure_ascii=False, indent=2)
        return StreamingResponse(
            iter([content]),
            media_type="application/json",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}_profiles.json"'},
        )

    elif format == "xlsx":
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{company_name} Profiles"[:31]  # Excel sheet name max 31 chars
        ws.append(EXPORT_HEADERS)
        for row in rows:
            ws.append([row.get(h, "") for h in EXPORT_HEADERS])
        _style_worksheet(ws)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}_profiles.xlsx"'},
        )

    elif format == "zip":
        import openpyxl

        folder = _sanitize_filename(company_name) or "export"
        photos_folder = f"{folder}/{folder}_photos"

        # Build xlsx in memory
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{company_name} Profiles"[:31]
        ws.append(EXPORT_HEADERS)
        for row in rows:
            ws.append([row.get(h, "") for h in EXPORT_HEADERS])
        _style_worksheet(ws)
        xlsx_buf = io.BytesIO()
        wb.save(xlsx_buf)
        xlsx_bytes = xlsx_buf.getvalue()

        # Build zip
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(f"{folder}/{folder}_profiles.xlsx", xlsx_bytes)

            for i, person in enumerate(people, 1):
                if not person.image_url:
                    continue
                photo_data = _download_photo(person.image_url)
                if not photo_data:
                    continue
                ext = _guess_extension(person.image_url, photo_data)
                first = _sanitize_filename(person.name.split()[0]) if person.name else "unknown"
                last = _sanitize_filename("_".join(person.name.split()[1:])) if person.name and len(person.name.split()) > 1 else ""
                photo_name = f"{i:03d}_{first}_{last}{ext}" if last else f"{i:03d}_{first}{ext}"
                zf.writestr(f"{photos_folder}/{photo_name}", photo_data)

        zip_buf.seek(0)
        return StreamingResponse(
            zip_buf,
            media_type="application/zip",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}_export.zip"'},
        )
